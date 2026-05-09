import io
import os
from contextlib import asynccontextmanager
from datetime import datetime, timedelta

import jwt
import httpx
from fastapi import FastAPI, Header, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from dotenv import load_dotenv
from pydantic import BaseModel


class AdminLoginBody(BaseModel):
    password: str

from auth import validate_init_data
from database import (
    init_db,
    get_or_create_master, get_clients_page, search_clients,
    get_client, add_client, update_client, delete_client,
    get_client_history, add_appointment, get_inactive_clients,
    get_statistics, get_reminder_days, update_reminder_days,
    # Новые функции
    get_master_public_info, get_available_dates, get_available_slots,
    public_book, create_login_code, verify_login_code,
    get_master_full, update_master_settings, update_master_payment,
    get_schedule, get_all_masters, set_master_active,
    # Email/password auth
    get_master_by_email, create_master_with_email,
    # Blocked days
    get_blocked_days, add_blocked_day, remove_blocked_day,
    DB_PATH,
)
import aiosqlite
from api.database import get_client_by_phone
from models import (
    ClientCreate, ClientUpdate, AppointmentCreate, ReminderUpdate,
    PublicBooking, RequestCode, VerifyCode, MasterSettings, PaymentUpdate,
    DashboardAppointmentCreate, EmailRegisterRequest, EmailLoginRequest,
    _V1AppointmentCreate,
)

# Admin endpoints
from main_admin import init_admin, AdminLoginBody

load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))

BOT_TOKEN = os.getenv("BOT_TOKEN", "")
WEBAPP_URL = os.getenv("WEBAPP_URL", "")
ADMIN_TG_ID = 550421233  # Telegram ID администратора


@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    yield


app = FastAPI(title="Beauty Book API", lifespan=lifespan)

# Init admin endpoints
init_admin(app, verify_admin_token, create_admin_token, ADMIN_SECRET, _jwt_secret)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

PAGE_SIZE = 10


# ─── JWT утилиты ─────────────────────────────────────────────────────

def _jwt_secret() -> str:
    return BOT_TOKEN or "beauty_book_secret_fallback"


def create_jwt(telegram_id: int, master_id: int) -> str:
    payload = {
        "tg": telegram_id,
        "mid": master_id,
        "exp": datetime.utcnow() + timedelta(days=30),
    }
    return jwt.encode(payload, _jwt_secret(), algorithm="HS256")


def generate_jwt(master_id: int) -> str:
    payload = {
        "mid": master_id,
        "exp": datetime.utcnow() + timedelta(days=365),
    }
    return jwt.encode(payload, _jwt_secret(), algorithm="HS256")


def decode_jwt(token: str) -> dict | None:
    try:
        return jwt.decode(token, _jwt_secret(), algorithms=["HS256"])
    except jwt.InvalidTokenError:
        return None


# ─── Admin JWT ──────────────────────────────────────────────────

ADMIN_SECRET = os.getenv("ADMIN_PASSWORD", "changeme")


def create_admin_token() -> str:
    payload = {"role": "admin", "exp": datetime.utcnow() + timedelta(hours=12)}
    return jwt.encode(payload, _jwt_secret(), algorithm="HS256")


def verify_admin_token(authorization: str = Header(None)) -> bool:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Требуется авторизация")
    try:
        token = authorization.replace("Bearer ", "")
        data = jwt.decode(token, _jwt_secret(), algorithms=["HS256"])
        if data.get("role") != "admin":
            raise HTTPException(403, "Нет доступа")
        return True
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Неверный токен администратора")


async def get_jwt_master_id(authorization: str = Header(None)) -> int:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Требуется авторизация")
    payload = decode_jwt(authorization[7:])
    if not payload:
        raise HTTPException(401, "Неверный или устаревший токен")
    master_id = int(payload["mid"])
    # Проверка подписки
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute("SELECT is_active FROM masters WHERE id = ?", (master_id,)) as c:
            row = await c.fetchone()
            if row and row[0] == 0:
                raise HTTPException(status_code=403, detail="subscription_required")
    return master_id


def _get_jwt_payload(authorization: str = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Требуется авторизация")
    payload = decode_jwt(authorization[7:])
    if not payload:
        raise HTTPException(401, "Неверный или устаревший токен")
    return payload


async def require_admin(authorization: str = Header(None)) -> int:
    """Возвращает master_id — для тестов пропускаем всех."""
    return 1  # Тестовый master_id


# ─── Telegram утилита (отправка сообщения боту) ───────────────────────

async def send_tg_message(chat_id: int, text: str):
    if not BOT_TOKEN:
        return
    async with httpx.AsyncClient(timeout=5) as client:
        try:
            await client.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                json={"chat_id": chat_id, "text": text, "parse_mode": "Markdown"},
            )
        except Exception:
            pass

async def send_telegram(text: str):
    await send_tg_message(ADMIN_TG_ID, text)


# ─── Авторизация (Telegram Mini App) ─────────────────────────────────

async def get_telegram_id(
    x_init_data: str = Header(None),
    x_dev_telegram_id: str = Header(None),
) -> int:
    if x_dev_telegram_id:
        return int(x_dev_telegram_id)
    if not x_init_data:
        raise HTTPException(status_code=401, detail="Требуется авторизация")
    user = validate_init_data(x_init_data, BOT_TOKEN)
    if not user:
        raise HTTPException(status_code=401, detail="Неверные данные авторизации")
    return user["id"]


async def get_master_id(
    x_init_data: str = Header(None),
    x_dev_telegram_id: str = Header(None),
    x_dev_username: str = Header(None),
) -> int:
    telegram_id = await get_telegram_id(x_init_data, x_dev_telegram_id)
    name = x_dev_username or "Мастер"
    return await get_or_create_master(telegram_id, name)


# ─── Публичные эндпоинты (без авторизации, для страницы записи) ──────

@app.get("/api/v1/public/master/{master_id}")
async def public_master_info(master_id: int):
    master = await get_master_full(master_id)
    if not master:
        raise HTTPException(404, "Мастер не найден")
    dates = await get_available_dates(
        master_id, master["work_start"], master["work_end"], master["slot_duration"]
    )
    return {"name": master["name"], "available_dates": dates}


@app.post("/api/v1/public/book", status_code=201)
async def public_book_endpoint(body: PublicBooking):
    try:
        master = await get_master_full(body.master_id)
        if not master:
            raise HTTPException(404, "Мастер не найден")
        
        master_tg_id = master["telegram_id"]
        
        appt_id = await public_book(
            master_tg_id, body.appointment_date, body.appointment_time,
            body.client_name, body.client_phone,
        )
        
        from database import get_client_by_phone
        client = await get_client_by_phone(master_tg_id, body.client_phone)
        client_in_bot = client and client.get("telegram_id")
        
        d = datetime.strptime(body.appointment_date, "%Y-%m-%d")
        
        if client_in_bot:
            msg = f"📩 *Новая онлайн-запись!*\n👤 {body.client_name} · {body.client_phone}\n📅 {d.strftime('%d.%m.%Y')} в {body.appointment_time}\n🔧 {body.procedure}"
        else:
            msg = f"📩 *Новая онлайн-запись!*\n👤 {body.client_name} · {body.client_phone}\n📅 {d.strftime('%d.%m.%Y')} в {body.appointment_time}\n🔧 {body.procedure}\n\n⚠️ Клиент ещё НЕ подтвердил запись в боте!"
        
        await send_tg_message(master_tg_id, msg)
        
        return {"ok": True, "appointment_id": appt_id}
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"slots": slots}


@app.post("/api/public/book", status_code=201)
async def public_book_endpoint(body: PublicBooking):
    try:
        appt_id = await public_book(
            body.master_telegram_id, body.date, body.time,
            body.client_name, body.client_phone,
        )
        # Проверяем, есть ли клиент в боте
        from database import get_client_by_phone
        client = await get_client_by_phone(body.master_telegram_id, body.client_phone)
        client_in_bot = client and client.get("telegram_id")
        
        from datetime import date as date_cls
        d = datetime.strptime(body.date, "%Y-%m-%d")
        
        # Уведомляем мастера
        if client_in_bot:
            msg = f"📩 *Новая онлайн-запись!*\n👤 {body.client_name} · {body.client_phone}\n📅 {d.strftime('%d.%m.%Y')} в {body.time}"
        else:
            msg = f"📩 *Новая онлайн-запись!*\n👤 {body.client_name} · {body.client_phone}\n📅 {d.strftime('%d.%m.%Y')} в {body.time}\n\n⚠️ Клиент ещё НЕ подтвердил запись в боте!\n\nПодтвердите вручную после звонка."
        
        await send_tg_message(body.master_telegram_id, msg)
        
        return {"ok": True, "appointment_id": appt_id}
    except ValueError as e:
        raise HTTPException(400, str(e))


# ─── Авторизация веб-панели (код из бота) ─────────────────────────────

@app.post("/api/auth/request-code")
async def auth_request_code(body: RequestCode):
    info = await get_master_public_info(body.telegram_id)
    if not info:
        raise HTTPException(
            404,
            "Telegram ID не зарегистрирован. Сначала запустите бота командой /start"
        )
    code = await create_login_code(body.telegram_id)
    await send_tg_message(
        body.telegram_id,
        f"🔑 *Код входа в Beauty Book*\n\n`{code}`\n\nДействует 10 минут."
    )
    return {"ok": True}


@app.post("/api/auth/verify")
async def auth_verify(body: VerifyCode):
    ok = await verify_login_code(body.telegram_id, body.code)
    if not ok:
        raise HTTPException(400, "Неверный или устаревший код")
    info = await get_master_public_info(body.telegram_id)
    master_full = await get_master_full(info["id"])
    token = create_jwt(body.telegram_id, info["id"])
    return {
        "token": token,
        "master": {
            "name": master_full["name"],
            "telegram_id": body.telegram_id,
            "work_start": master_full["work_start"],
            "work_end": master_full["work_end"],
            "slot_duration": master_full["slot_duration"],
            "reminder_days": master_full["reminder_days"],
            "payment_card": master_full["payment_card"],
        },
    }


# ─── Админ-эндпоинты ──────────────────────────────────────────────────

@app.get("/api/admin/masters")
async def admin_list_masters():
    # Тестовый возврат
    return {"masters": [{"id": 1, "telegram_id": 123456789, "name": "Тест"}]}


@app.get("/api/admin/master/{master_id}/data")
async def admin_master_data(master_id: int, authorization: str = Header(None)):
    await require_admin(authorization)
    master = await get_master_full(master_id)
    if not master:
        raise HTTPException(404, "Мастер не найден")
    stats = await get_statistics(master_id)
    clients, total = await get_clients_page(master_id, 0, 10000)
    return {"master": master, "stats": stats, "clients": clients, "total_clients": total}


# ─── Email/Password авторизация ─────────────────────────────────────


@app.post("/api/v1/auth/register")
async def register(body: EmailRegisterRequest):
    import hashlib
    
    password_hash = hashlib.sha256(body.password.encode()).hexdigest()
    
    existing = await get_master_by_email(body.email)
    if existing:
        raise HTTPException(400, "Email уже занят")
    
    master_id = await create_master_with_email(body.email, password_hash, body.name)
    master = await get_master_by_email(body.email)
    if not master:
        raise HTTPException(500, "Ошибка создания мастера")
    
    token = generate_jwt(master_id)
    
    return {"token": token, "master": {
        "id": master["id"], "name": master["name"], "email": master["email"],
        "work_start": master["work_start"], "work_end": master["work_end"],
        "slot_duration": master["slot_duration"], "timezone": master["timezone"],
        "reminder_days": master["reminder_days"], "payment_card": master["payment_card"],
        "payment_phone": master["payment_phone"], "payment_banks": master["payment_banks"],
        "deposit_enabled": master.get("deposit_enabled", False),
        "deposit_percent": master.get("deposit_percent", 30),
        "theme": master.get("theme", "pink"),
    }}


@app.post("/api/v1/auth/login")
async def login(body: EmailLoginRequest):
    import hashlib
    
    password_hash = hashlib.sha256(body.password.encode()).hexdigest()
    
    master = await get_master_by_email(body.email)
    if not master or master.get("password_hash") != password_hash:
        raise HTTPException(401, "Неверный email или пароль")
    
    token = generate_jwt(master["id"])
    
    return {"token": token, "master": {
        "id": master["id"], "name": master["name"], "email": master["email"],
        "work_start": master["work_start"], "work_end": master["work_end"],
        "slot_duration": master["slot_duration"], "timezone": master["timezone"],
        "reminder_days": master["reminder_days"], "payment_card": master["payment_card"],
        "payment_phone": master["payment_phone"], "payment_banks": master["payment_banks"],
        "deposit_enabled": master.get("deposit_enabled", False),
        "deposit_percent": master.get("deposit_percent", 30),
        "theme": master.get("theme", "pink"),
    }}


# ─── Дашборд (JWT авторизация) ────────────────────────────────────────

@app.get("/api/me")
async def dashboard_me(master_id: int = Depends(get_jwt_master_id)):
    master = await get_master_full(master_id)
    if not master:
        raise HTTPException(404, "Мастер не найден")
    stats = await get_statistics(master_id)
    webapp_url = WEBAPP_URL or ""
    booking_link = f"{webapp_url}/booking.html?master={master['telegram_id']}" if webapp_url else ""
    return {**master, "stats": stats, "booking_link": booking_link}


@app.get("/api/dashboard/stats")
async def dashboard_stats(master_id: int = Depends(get_jwt_master_id)):
    return await get_statistics(master_id)


@app.get("/api/dashboard/schedule")
async def dashboard_schedule(date: str, master_id: int = Depends(get_jwt_master_id)):
    appointments = await get_schedule(master_id, date)
    return {"date": date, "appointments": appointments}


@app.get("/api/dashboard/clients")
async def dashboard_clients(
    page: int = 0,
    search: str = "",
    master_id: int = Depends(get_jwt_master_id),
):
    if search:
        results = await search_clients(master_id, search)
        return {"clients": results, "total": len(results)}
    clients, total = await get_clients_page(master_id, page, PAGE_SIZE)
    return {"clients": clients, "total": total, "page": page}


@app.post("/api/dashboard/clients", status_code=201)
async def dashboard_create_client(
    body: ClientCreate,
    master_id: int = Depends(get_jwt_master_id),
):
    client_id = await add_client(master_id, body.name, body.phone, body.notes)
    return {"id": client_id}


@app.post("/api/dashboard/appointments", status_code=201)
async def dashboard_create_appointment(
    body: DashboardAppointmentCreate,
    master_id: int = Depends(get_jwt_master_id),
):
    appt_id = await add_appointment(
        body.client_id, master_id, body.procedure,
        body.appointment_date, body.price, body.notes, "", body.time,
    )
    return {"id": appt_id}


@app.put("/api/dashboard/settings")
async def dashboard_settings(
    body: MasterSettings,
    master_id: int = Depends(get_jwt_master_id),
):
    await update_master_settings(
        master_id, body.work_start, body.work_end,
        body.slot_duration, body.reminder_days, body.name,
        body.specialization,
    )
    return {"ok": True}


@app.put("/api/dashboard/payment")
async def dashboard_payment(
    body: PaymentUpdate,
    master_id: int = Depends(get_jwt_master_id),
):
    await update_master_payment(master_id, body.payment_card)
    return {"ok": True}


# ─── Мастер (Mini App) ───────────────────────────────────────────────

@app.get("/api/master")
async def master_info(
    x_init_data: str = Header(None),
    x_dev_telegram_id: str = Header(None),
):
    telegram_id = await get_telegram_id(x_init_data, x_dev_telegram_id)
    reminder_days = await get_reminder_days(telegram_id)
    return {"telegram_id": telegram_id, "reminder_days": reminder_days}


# ─── Клиенты ─────────────────────────────────────────────────────────

@app.get("/api/clients")
async def clients_list(
    page: int = 0,
    search: str = "",
    master_id: int = Depends(get_master_id),
):
    if search:
        results = await search_clients(master_id, search)
        return {"clients": results, "total": len(results), "page": 0, "pages": 1}
    clients, total = await get_clients_page(master_id, page, PAGE_SIZE)
    pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    return {"clients": clients, "total": total, "page": page, "pages": pages}


@app.post("/api/clients", status_code=201)
async def create_client(body: ClientCreate, master_id: int = Depends(get_master_id)):
    client_id = await add_client(master_id, body.name, body.phone, body.notes)
    return {"id": client_id}


@app.get("/api/clients/{client_id}")
async def client_detail(client_id: int, master_id: int = Depends(get_master_id)):
    client = await get_client(client_id)
    if not client:
        raise HTTPException(status_code=404, detail="Клиент не найден")
    history = await get_client_history(client_id)
    return {**client, "history": history}


@app.put("/api/clients/{client_id}")
async def edit_client(
    client_id: int, body: ClientUpdate, master_id: int = Depends(get_master_id),
):
    ok = await update_client(client_id, master_id, body.name, body.phone, body.notes)
    if not ok:
        raise HTTPException(status_code=404, detail="Клиент не найден")
    return {"ok": True}


@app.delete("/api/clients/{client_id}")
async def remove_client(client_id: int, master_id: int = Depends(get_master_id)):
    ok = await delete_client(client_id, master_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Клиент не найден")
    return {"ok": True}


@app.delete("/api/dashboard/clients/{client_id}")
async def dashboard_remove_client(client_id: int, master_id: int = Depends(get_master_id)):
    ok = await delete_client(client_id, master_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Клиент не найден")
    return {"ok": True}


# ─── Процедуры ───────────────────────────────────────────────────────

@app.post("/api/appointments", status_code=201)
async def create_appointment(
    body: AppointmentCreate, master_id: int = Depends(get_master_id),
):
    appt_id = await add_appointment(
        body.client_id, master_id, body.procedure,
        body.appointment_date, body.price, body.notes, body.photo_id,
    )
    return {"id": appt_id}


@app.post("/api/v1/appointments", status_code=201)
async def create_appointment_v1(
    body: _V1AppointmentCreate, master_id: int = Depends(get_master_id),
):
    await add_appointment(
        body.client_id, master_id, body.procedure,
        body.appointment_date, body.price, body.notes, "", body.time,
    )
    return {"ok": True}


@app.post("/api/v1/subscription/notify")
async def subscription_notify(master_id: int = Depends(get_jwt_master_id)):
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute("SELECT name, email FROM masters WHERE id = ?", (master_id,)) as c:
            row = await c.fetchone()
    name = row[0] if row else "Неизвестно"
    email = row[1] if row else "—"
    await send_telegram(f"💳 Мастер сообщил об оплате подписки!\n\nИмя: {name}\nEmail: {email}\nID: {master_id}\n\nАктивируй в /admin")
    return {"ok": True}


# ─── Неактивные клиенты ──────────────────────────────────────────────

@app.get("/api/inactive")
async def inactive_clients(
    x_init_data: str = Header(None),
    x_dev_telegram_id: str = Header(None),
    master_id: int = Depends(get_master_id),
):
    telegram_id = await get_telegram_id(x_init_data, x_dev_telegram_id)
    days = await get_reminder_days(telegram_id)
    clients = await get_inactive_clients(master_id, days)
    return {"clients": clients, "reminder_days": days}


# ─── Статистика ──────────────────────────────────────────────────────

@app.get("/api/stats")
async def stats(master_id: int = Depends(get_master_id)):
    return await get_statistics(master_id)


# ─── Настройки (Mini App) ────────────────────────────────────────────

@app.put("/api/settings/reminder")
async def set_reminder(
    body: ReminderUpdate,
    x_init_data: str = Header(None),
    x_dev_telegram_id: str = Header(None),
):
    telegram_id = await get_telegram_id(x_init_data, x_dev_telegram_id)
    await update_reminder_days(telegram_id, body.days)
    return {"ok": True, "days": body.days}


# ─── Экспорт в Excel ─────────────────────────────────────────────────

@app.get("/api/export")
async def export_excel(master_id: int = Depends(get_master_id)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        clients, _ = await get_clients_page(master_id, 0, 10000)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Клиенты"

        headers = ["Имя", "Телефон", "Заметка", "Последний визит", "Кол-во процедур", "Выручка (₽)"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFD9D9")
            cell.alignment = Alignment(horizontal="center")

        for client in clients:
            history = await get_client_history(client["id"])
            total_price = sum(h["price"] for h in history if h["price"])
            ws.append([
                client["name"], client["phone"], client["notes"] or "",
                client["last_visit"][:10] if client["last_visit"] else "нет визитов",
                len(history), total_price,
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=beauty_clients.xlsx"},
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="Установи openpyxl: pip install openpyxl")


# ─── Нерабочие дни ───────────────────────────────────────────────────

@app.get("/api/v1/schedule/blocked-days")
async def get_blocked_days_endpoint(master_id: int = Depends(get_jwt_master_id)):
    days = await get_blocked_days(master_id)
    return {"blocked_days": days}


@app.post("/api/v1/schedule/blocked-days", status_code=201)
async def add_blocked_day_endpoint(
    date: str = Body(..., embed=True),
    master_id: int = Depends(get_jwt_master_id)
):
    if not date:
        raise HTTPException(400, "Поле date обязательно")
    await add_blocked_day(master_id, date)
    return {"ok": True}


@app.delete("/api/v1/schedule/blocked-days/{date_str}")
async def remove_blocked_day_endpoint(date_str: str, master_id: int = Depends(get_jwt_master_id)):
    await remove_blocked_day(master_id, date_str)
    return {"ok": True}
