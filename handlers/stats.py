import io
from datetime import date, timedelta
from aiogram import Router, F
from aiogram.types import CallbackQuery, BufferedInputFile

from database import (
    get_or_create_master, get_statistics, get_clients, get_client_history,
    get_earnings_by_period, get_earnings_by_service, get_earnings_by_client,
    get_earnings_by_day,
)
from keyboards import stats_menu_keyboard, back_to_menu

router = Router()

MONTHS_RU = {
    1: "Янв", 2: "Фев", 3: "Мар", 4: "Апр",
    5: "Май", 6: "Июн", 7: "Июл", 8: "Авг",
    9: "Сен", 10: "Окт", 11: "Ноя", 12: "Дек",
}


def _bar(value: int, max_value: int, width: int = 12) -> str:
    if max_value == 0:
        return ""
    filled = round(value / max_value * width)
    return "█" * filled + "░" * (width - filled)


@router.callback_query(F.data == "stats")
async def cb_stats(callback: CallbackQuery):
    master_id = await get_or_create_master(callback.from_user.id, callback.from_user.full_name)
    stats = await get_statistics(master_id)

    text = "📊 *Статистика*\n\n"
    text += f"👥 Всего клиентов: *{stats['total_clients']}*\n"
    text += f"📅 Всего процедур: *{stats['total_appointments']}*\n"
    text += f"💰 Общая выручка: *{stats['total_earnings']}₽*\n"
    text += f"📆 За этот месяц: *{stats['month_earnings']}₽*\n"

    if stats["top_procedures"]:
        text += "\n🏆 *Топ услуги:*\n"
        medals = ["🥇", "🥈", "🥉"]
        for i, p in enumerate(stats["top_procedures"][:3]):
            proc = p["procedure"] if isinstance(p, dict) else p[0]
            cnt = p["count"] if isinstance(p, dict) else p[1]
            text += f"{medals[i]} {proc} — {cnt} раз\n"

    text += "\n_Выбери вид статистики:_"
    await callback.message.edit_text(text, reply_markup=stats_menu_keyboard(), parse_mode="Markdown")
    await callback.answer()


@router.callback_query(F.data.startswith("stats_period:"))
async def cb_stats_period(callback: CallbackQuery):
    period = callback.data.split(":")[1]
    master_id = await get_or_create_master(callback.from_user.id, callback.from_user.full_name)

    today = date.today()
    if period == "day":
        date_from = date_to = today.isoformat()
        label = f"Сегодня, {today.strftime('%d.%m.%Y')}"
    elif period == "week":
        date_from = (today - timedelta(days=6)).isoformat()
        date_to = today.isoformat()
        label = f"Последние 7 дней"
    elif period == "month":
        date_from = today.replace(day=1).isoformat()
        date_to = today.isoformat()
        label = f"{MONTHS_RU[today.month]} {today.year}"
    else:  # year
        date_from = today.replace(month=1, day=1).isoformat()
        date_to = today.isoformat()
        label = f"{today.year} год"

    data = await get_earnings_by_period(master_id, date_from, date_to)
    text = (
        f"📊 *Заработок — {label}*\n\n"
        f"📅 Процедур: *{data['total_appointments']}*\n"
        f"💰 Выручка: *{data['total_earnings']}₽*\n"
    )

    by_svc = await get_earnings_by_service(master_id, date_from, date_to)
    if by_svc:
        text += "\n💅 *По услугам:*\n"
        for proc, cnt, total in by_svc[:8]:
            text += f"• {proc}: {cnt} раз · *{total}₽*\n"

    await callback.message.edit_text(text, reply_markup=stats_menu_keyboard(), parse_mode="Markdown")
    await callback.answer()


@router.callback_query(F.data == "stats_by_service")
async def cb_stats_by_service(callback: CallbackQuery):
    master_id = await get_or_create_master(callback.from_user.id, callback.from_user.full_name)
    rows = await get_earnings_by_service(master_id)

    if not rows:
        await callback.answer("Нет данных", show_alert=True)
        return

    max_total = max(r[2] for r in rows) or 1
    text = "💅 *Доход по услугам (за всё время)*\n\n"
    for proc, cnt, total in rows[:15]:
        bar = _bar(total, max_total)
        text += f"`{bar}` *{total}₽*\n{proc} — {cnt} раз\n\n"

    await callback.message.edit_text(text, reply_markup=stats_menu_keyboard(), parse_mode="Markdown")
    await callback.answer()


@router.callback_query(F.data == "stats_by_client")
async def cb_stats_by_client(callback: CallbackQuery):
    master_id = await get_or_create_master(callback.from_user.id, callback.from_user.full_name)
    rows = await get_earnings_by_client(master_id)

    if not rows:
        await callback.answer("Нет данных", show_alert=True)
        return

    text = "👥 *Доход по клиентам (топ-20)*\n\n"
    medals = ["🥇", "🥈", "🥉"]
    for i, (name, cnt, total) in enumerate(rows):
        icon = medals[i] if i < 3 else f"{i + 1}."
        text += f"{icon} *{name}* — {total}₽ · {cnt} визитов\n"

    await callback.message.edit_text(text, reply_markup=stats_menu_keyboard(), parse_mode="Markdown")
    await callback.answer()


@router.callback_query(F.data == "stats_chart")
async def cb_stats_chart(callback: CallbackQuery):
    master_id = await get_or_create_master(callback.from_user.id, callback.from_user.full_name)
    rows = await get_earnings_by_day(master_id, 30)

    if not rows:
        await callback.answer("Нет данных за последние 30 дней", show_alert=True)
        return

    max_val = max(r[1] for r in rows) or 1
    text = "📈 *График дохода — последние 30 дней*\n\n"
    for day_str, total in rows:
        d = date.fromisoformat(day_str)
        label = f"{d.day:02d} {MONTHS_RU[d.month]}"
        bar = _bar(total, max_val, width=10)
        text += f"`{label}` {bar} {total}₽\n"

    await callback.message.edit_text(text, reply_markup=stats_menu_keyboard(), parse_mode="Markdown")
    await callback.answer()


@router.callback_query(F.data == "export_excel")
async def cb_export_excel(callback: CallbackQuery):
    await callback.answer("Генерирую файл...")
    master_id = await get_or_create_master(callback.from_user.id, callback.from_user.full_name)
    clients = await get_clients(master_id)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()

        # Лист 1: Клиенты
        ws = wb.active
        ws.title = "Клиенты"
        headers = ["Имя", "Телефон", "Заметка", "Последний визит", "Кол-во процедур", "Выручка (₽)"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFD9D9")
            cell.alignment = Alignment(horizontal="center")

        for client in clients:
            cid, name, phone, notes, last_visit = client
            history = await get_client_history(cid)
            total_price = sum(h[2] for h in history if h[2])
            ws.append([
                name, phone, notes or "",
                last_visit[:10] if last_visit else "нет визитов",
                len(history), total_price,
            ])

        # Лист 2: По услугам
        ws2 = wb.create_sheet("По услугам")
        ws2.append(["Услуга", "Процедур", "Выручка (₽)"])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFD9D9")
        by_svc = await get_earnings_by_service(master_id)
        for proc, cnt, total in by_svc:
            ws2.append([proc, cnt, total])

        # Лист 3: По клиентам
        ws3 = wb.create_sheet("По клиентам")
        ws3.append(["Клиент", "Визитов", "Выручка (₽)"])
        for cell in ws3[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFD9D9")
        by_client = await get_earnings_by_client(master_id)
        for name, cnt, total in by_client:
            ws3.append([name, cnt, total])

        for sheet in [ws, ws2, ws3]:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        file = BufferedInputFile(buf.read(), filename="beauty_clients.xlsx")
        await callback.message.answer_document(file, caption="📥 База выгружена (клиенты, услуги, статистика)")

    except ImportError:
        await callback.message.answer(
            "❌ Для экспорта установи библиотеку:\n`pip install openpyxl`",
            parse_mode="Markdown",
            reply_markup=back_to_menu()
        )
